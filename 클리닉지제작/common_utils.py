import yaml
from typing import Dict, Any
from openpyxl import load_workbook
import os
from pathlib import Path

class PathManager:
    def __init__(self):
        # 현재 실행 중인 파일의 디렉토리(클리닉지제작)를 기준으로 경로 설정
        self.base_dir = Path(__file__).parent
        self._create_directory_structure()

    def _create_directory_structure(self):
        """필요한 디렉토리 구조 생성"""
        (self.base_dir / "output").mkdir(exist_ok=True)
        (self.base_dir / "output" / "generated_problems").mkdir(exist_ok=True)
        (self.base_dir / "output" / "validation_results").mkdir(exist_ok=True)

    def get_absolute_path(self, relative_path: str) -> Path:
        """상대 경로를 절대 경로로 변환"""
        return self.base_dir / relative_path

    def get_config_path(self) -> Path:
        """config.yaml 파일의 경로"""
        return self.base_dir / "config.yaml"

class ConfigManager:
    def __init__(self, config_path: str = None):
        self.path_manager = PathManager()
        self.config_path = config_path or self.path_manager.get_config_path()
        self.config = self._load_config()

    def _load_config(self) -> Dict[str, Any]:
        with open(self.config_path, 'r', encoding='utf-8') as file:
            return yaml.safe_load(file)

    def get_gpt_settings(self, purpose: str) -> Dict[str, Any]:
        return self.config['gpt_settings'].get(purpose, {})

    def get_prompt(self, prompt_type: str) -> str:
        return self.config['prompts'].get(prompt_type, '')

    def get_api_key(self) -> str:
        return self.config['api'].get('key', '')

    def get_paths(self) -> Dict[str, Path]:
        paths = self.config.get('paths', {})
        return {
            'excel_file': self.path_manager.get_absolute_path(paths['excel_file']),
            'output_folder': self.path_manager.get_absolute_path(paths['output_folder']),
            'generated_problems_folder': self.path_manager.get_absolute_path(paths['generated_problems_folder']),
            'validation_results_folder': self.path_manager.get_absolute_path(paths['validation_results_folder'])
        }

    def get_excel_settings(self) -> Dict[str, Any]:
        return self.config.get('excel_settings', {})
        
    def get_output_settings(self) -> Dict[str, Any]:
        return self.config.get('output_settings', {})

class FileNameManager:
    def __init__(self, config_manager: ConfigManager):
        self.config_manager = config_manager
        self.paths = config_manager.get_paths()
        self.excel_settings = config_manager.get_excel_settings()
        self.output_settings = config_manager.get_output_settings()
        self._base_filename = None

    def get_base_filename(self) -> str:
        """엑셀에서 기본 파일명을 가져옴"""
        if self._base_filename is None:
            try:
                wb = load_workbook(self.paths['excel_file'])
                ws = wb.active
                cell_address = self.excel_settings.get('filename_cell', 'E1')
                self._base_filename = ws[cell_address].value or "default"
                wb.close()
            except Exception as e:
                print(f"Error reading filename from Excel: {str(e)}")
                self._base_filename = "default"
        return self._base_filename

    def get_json_path(self) -> Path:
        """JSON 파일 경로 생성"""
        settings = self.output_settings
        filename = (f"{settings.get('filename_prefix', '')}"
                   f"{self.get_base_filename()}"
                   f"{settings.get('filename_suffix', '')}"
                   f"{settings.get('json_extension', '.json')}")
        return self.paths['generated_problems_folder'] / filename

    def get_word_path(self) -> Path:
        """Word 파일 경로 생성"""
        settings = self.output_settings
        filename = (f"{settings.get('filename_prefix', '')}"
                   f"{self.get_base_filename()}"
                   f"{settings.get('filename_suffix', '')}"
                   f"{settings.get('word_extension', '.docx')}")
        return self.paths['generated_problems_folder'] / filename

    def get_validation_path(self) -> Path:
        """검증 결과 파일 경로 생성"""
        settings = self.output_settings
        filename = (f"{settings.get('filename_prefix', '')}"
                   f"{self.get_base_filename()}"
                   f"{settings.get('validation_extension', '_validation.json')}")
        return self.paths['validation_results_folder'] / filename
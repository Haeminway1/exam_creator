import json
import logging
from openai import OpenAI
from openpyxl import load_workbook
import os
from pathlib import Path
from typing import Dict, Any, List
from common_utils import ConfigManager, FileNameManager, PathManager

class ProblemGenerator:
    def __init__(self, config_path: str = None):
        self.path_manager = PathManager()
        self.config_manager = ConfigManager(config_path)
        self.file_manager = FileNameManager(self.config_manager)
        self.gpt_settings = self.config_manager.get_gpt_settings('problem_generation')
        self.prompt_template = self.config_manager.get_prompt('problem_generation')
        
        self.paths = self.config_manager.get_paths()
        self.excel_path = self.paths['excel_file']
        self.output_folder = self.paths['generated_problems_folder']
        
        self.api_key = self.config_manager.get_api_key()
        self.client = OpenAI(api_key=self.api_key)
        self.keywords = []
        self.logger = self._setup_logger()
        self.excel_settings = self.config_manager.get_excel_settings()

    def _setup_logger(self) -> logging.Logger:
        logger = logging.getLogger('problem_generator')
        logger.setLevel(logging.INFO)
        
        # 로그 파일을 output_folder에 저장
        log_file = self.output_folder / 'problem_generator.log'
        os.makedirs(log_file.parent, exist_ok=True)
        
        file_handler = logging.FileHandler(
            str(log_file),
            encoding='utf-8'
        )
        file_handler.setLevel(logging.INFO)
        
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)
        
        return logger

    def load_keywords_from_excel(self) -> None:
        try:
            wb = load_workbook(str(self.excel_path))
            ws = wb.active
            self.keywords = []
            
            settings = self.excel_settings
            start_row = settings.get('start_row', 2)
            column_range = settings.get('column_range', ['C', 'D', 'E'])
            max_row = ws.max_row
            
            # 첫 번째 열만 필수로 검사
            first_column = column_range[0]
            if not any(ws[f"{first_column}{row}"].value for row in range(start_row, max_row + 1)):
                self.logger.error(f"Required column {first_column} is empty")
                raise ValueError(f"Required column {first_column} is empty")
            
            for row in ws.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
                keyword = {}
                for i, col in enumerate(column_range):
                    col_index = ord(col.upper()) - ord('A')
                    if i == 0:  # 첫 번째 열은 필수
                        value = row[col_index] if len(row) > col_index and row[col_index] is not None else ''
                        if not value:  # 첫 번째 열이 비어있으면 해당 행 스킵
                            continue
                    else:  # 나머지 열은 선택
                        value = row[col_index] if len(row) > col_index and row[col_index] is not None else ''
                    keyword[f'열{i+1}'] = value
                
                if keyword.get('열1'):  # 첫 번째 열이 있는 경우만 추가
                    self.keywords.append(keyword)
                    
            self.logger.info(f"Successfully loaded {len(self.keywords)} keywords from Excel")
            wb.close()
            
        except Exception as e:
            self.logger.error(f"Error loading keywords from Excel: {str(e)}")
            raise

    def generate_prompt(self, keyword: Dict[str, str]) -> str:
        """프롬프트 생성 - 추가 열이 있는 경우만 포함"""
        base_prompt = self.prompt_template.format(본문=keyword['열1'])
        
        # 추가 열 데이터가 있는 경우 프롬프트에 추가
        additional_info = []
        if keyword.get('열2'):
            additional_info.append(f"추가 정보 1: {keyword['열2']}")
        if keyword.get('열3'):
            additional_info.append(f"추가 정보 2: {keyword['열3']}")
        
        if additional_info:
            base_prompt += "\n추가 컨텍스트:\n" + "\n".join(additional_info)
        
        return base_prompt
    
    def chat_with_gpt_and_collect(self, prompt: str) -> Dict:
        try:
            # 시스템 메시지에 JSON 응답 형식을 명시적으로 포함
            system_message = (
                f"{self.gpt_settings['system_prompt']} "
                "응답은 반드시 다음과 같은 JSON 형식으로 제공해주세요: "
                '{"문제1": {"문제": "문제1 내용", "선지": ["1.", "2.", "3.", "4.", "5."]}, '
                '"문제2": {"문제": "문제2 내용", "선지": ["1.", "2.", "3.", "4.", "5."]}, '
                '"정답": ["정답1", "정답2"]}'
            )

            response = self.client.chat.completions.create(
                model=self.gpt_settings['model'],
                messages=[
                    {"role": "system", "content": system_message},
                    {"role": "user", "content": prompt}
                ],
                temperature=float(self.gpt_settings['temperature']),
                max_tokens=int(self.gpt_settings['max_tokens']),
                top_p=float(self.gpt_settings['top_p']),
                frequency_penalty=float(self.gpt_settings['frequency_penalty']),
                presence_penalty=float(self.gpt_settings['presence_penalty']),
                response_format={"type": "json_object"}
            )
            
            return json.loads(response.choices[0].message.content.strip())
            
        except Exception as e:
            self.logger.error(f"Error in chat_with_gpt_and_collect: {str(e)}")
            return {"error": str(e)}

    def create_initial_json(self) -> List[Dict[str, Any]]:
        initial_data = []
        for i, keyword in enumerate(self.keywords, start=1):
            initial_data.append({
                "본문번호": i,
                "본문": keyword['열1'],
                "문제들": []
            })
        return initial_data

    def process_gpt_response(self, response: Dict[str, Any], item: Dict[str, Any], index: int) -> None:
        """GPT 응답 처리 및 문제 데이터 구조화"""
        try:
            if "error" not in response:
                # 문제 및 답안 처리
                for i, (key, value) in enumerate(response.items(), start=1):
                    if isinstance(value, dict) and "문제" in value and "선지" in value:
                        item["문제들"].append({
                            "문제번호": f"{index+1}-{i}",
                            "문제": value["문제"],
                            "선지": value["선지"]
                        })
                
                # 정답 처리
                if "정답" in response:
                    item["정답"] = response["정답"]
                else:
                    self.logger.warning(f"No answer found for problem set {index+1}")
            else:
                self.logger.error(f"Error in response for item {index + 1}: {response['error']}")
                
        except Exception as e:
            self.logger.error(f"Error processing GPT response: {str(e)}")
            raise

    def generate_draft(self) -> Dict[str, str]:
        try:
            self.load_keywords_from_excel()
            initial_data = self.create_initial_json()
            
            max_retries = 3
            total_keywords = len(self.keywords)
            
            for retry in range(max_retries):
                try:
                    for index, item in enumerate(initial_data):
                        self.logger.info(f"Processing item {index + 1} of {total_keywords}")
                        prompt = self.generate_prompt({"열1": item["본문"]})
                        self.logger.info(f"Generated prompt for item {index + 1}")
                        
                        response = self.chat_with_gpt_and_collect(prompt)
                        self.logger.info(f"Received response for item {index + 1}")
                        
                        self.process_gpt_response(response, item, index)
                        
                        if "error" in response:
                            continue
                    
                    self.save_responses_to_json(initial_data)
                    return {
                        "message": "Drafts generated and compiled successfully",
                        "total_passages": total_keywords,
                        "total_problems": sum(len(item["문제들"]) for item in initial_data)
                    }
                    
                except Exception as e:
                    self.logger.error(f"Error in generate_draft (attempt {retry+1}/{max_retries}): {str(e)}")
                    if retry == max_retries - 1:
                        return {"error": f"Maximum retries reached. Failed to generate draft: {str(e)}"}
            
            return {"error": "An unexpected error occurred while generating the draft"}
            
        except Exception as e:
            self.logger.error(f"Critical error in generate_draft: {str(e)}")
            return {"error": f"Critical error: {str(e)}"}

    def save_responses_to_json(self, responses: List[Dict[str, Any]]) -> None:
        try:
            file_path = self.file_manager.get_json_path()
            os.makedirs(file_path.parent, exist_ok=True)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(responses, f, ensure_ascii=False, indent=2)
            self.logger.info(f"Successfully saved responses to {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error saving responses to JSON: {str(e)}")
            raise

    def validate_excel_format(self) -> bool:
        """Excel 파일 형식 검증 - 첫 번째 열만 필수로 확인"""
        try:
            if not self.excel_path.exists():
                self.logger.error(f"Excel file not found: {self.excel_path}")
                return False

            wb = load_workbook(str(self.excel_path))
            ws = wb.active
            
            # 파일명 셀 확인
            filename_cell = self.excel_settings.get('filename_cell', 'E1')
            if not ws[filename_cell].value:
                self.logger.error(f"Filename cell {filename_cell} is empty")
                return False

            # 첫 번째 열만 필수로 확인
            first_column = self.excel_settings.get('column_range', ['C', 'D', 'E'])[0]
            if not any(ws[f"{first_column}{row}"].value for row in range(2, ws.max_row + 1)):
                self.logger.error(f"Required column {first_column} is empty")
                return False

            wb.close()
            return True
            
        except Exception as e:
            self.logger.error(f"Error validating Excel format: {str(e)}")
            return False

def main():
    try:
        # 설정 파일 경로를 자동으로 찾음
        generator = ProblemGenerator()
        
        # Excel 형식 검증
        if not generator.validate_excel_format():
            print("Error: Invalid Excel format. Please check the log file for details.")
            return
        
        result = generator.generate_draft()
        
        if "error" not in result:
            print(f"\nGeneration Summary:")
            print(f"Total Passages: {result['total_passages']}")
            print(f"Total Problems: {result['total_problems']}")
            print(f"Status: {result['message']}")
        else:
            print(f"Error: {result['error']}")
        
    except Exception as e:
        print(f"Error in main: {str(e)}")

if __name__ == "__main__":
    main()